import sys
from pathlib import Path
from dotenv import load_dotenv
import certifi
import requests
import os
import time
import shutil
import zipfile
from tqdm import tqdm
import pandas as pd
import traceback
from datetime import date,datetime
from dateutil.relativedelta import relativedelta
import locale
import calendar
import os
import json
import re
from openpyxl.styles import PatternFill,Alignment

locale.setlocale(locale.LC_TIME, "Russian_Russia")

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = Path(sys.executable).parent
    else:
        base_path = Path(__file__).parent
    return base_path / relative_path


dotenv_path = resource_path(".env")
data_path = resource_path("DATA")
output_path = resource_path("OUTPUT")
load_dotenv(dotenv_path=dotenv_path)
TOKEN_YANDEX = os.getenv("TOKEN_YANDEX")
BUSINESSID = os.getenv("businessId")
session = requests.Session()
session.verify = certifi.where()
session.headers.update({
    "Api-Key": TOKEN_YANDEX,
    "Content-Type": "application/json",
    "Accept": "application/json"
})

MENEGERS = ["Косых Дмитрий", "Троянва Олеся","Бирюков Алексей"]

def create_yandex_report(date_from,date_to):
    url_create = "https://api.partner.market.yandex.ru/reports/united-orders/generate"
    params = {"format":"CSV","ReportLanguageType":"RU"}

    payload = {
        "businessId": BUSINESSID,
        "dateFrom": date_from,
        "dateTo": date_to,
        "campaignIds":[22600620,148607624]
    }
    response = session.post(url=url_create,params=params,json = payload)
    if response.status_code == 200:
        data = response.json()
        report_id = data.get("result",{}).get("reportId")
        print("Отчёт YandexMarket успешно создан. Информация:")
        print(f"ID отчета: {report_id}")
        return report_id
    else:
        print(f"Ошибка запроса: {response.status_code}")
        print(response.text)

def get_yandex_report(report_id,mask = True):
    url_get = f"https://api.partner.market.yandex.ru/reports/info/{report_id}"
    save_path = "report.zip"
    extract_path = "unzipped_report"
    while True:
        response = session.get(url=url_get)
        if response.status_code != 200:
            raise RuntimeError(f"Ошибка создания отчета: {response.status_code} {response.text}")

        data = response.json()
        result = data.get("result",{})
        status = result.get("status")
        if status == "DONE":
            file_link = result.get("file")
            print("Отчет YandexMarket готов. Скачивание файла...")
            with session.get(file_link, stream=True) as r:
                r.raise_for_status()
                total_size = int(r.headers.get('content-length', 0))
                block_size = 1024
                with open(save_path, 'wb') as f, tqdm(
                    desc="Скачивание",
                    total=total_size,
                    unit='iB',
                    unit_scale=True,
                    unit_divisor=1024
                ) as bar:
                    for data in r.iter_content(block_size):
                        f.write(data)
                        bar.update(len(data))
            if zipfile.is_zipfile(save_path):
                with zipfile.ZipFile(save_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)
                print(f"Файл успешно распакован в папку: {extract_path}")
                if os.path.exists(save_path):
                    os.remove(save_path)
                break
            else:
                print("Скачанный файл — не архив ZIP")
                break
        elif status == "FAILED":
            print(f"Ошибка генерации отчета: {status}")
            break
        else:
            print(f"Отчёт ещё не готов. Ожидаем 10 секунд...")
            time.sleep(10)
    if mask:
        report_path = f"{extract_path}\\orders_and_offers_transactions.csv"
        df = pd.read_csv(report_path,low_memory=False)
        extract_path = Path("unzipped_report")
        if extract_path.exists():
            shutil.rmtree(extract_path)
        return df
    else:
        report_pay_path = f"{extract_path}\\netting_report_accruals.csv"
        report_return_path = f"{extract_path}\\netting_report_returns_and_compensations.csv"
        df_payments = pd.read_csv(report_pay_path,low_memory=False)
        df_payments.rename(columns={"ORDER_ID":"Номер по данным клиента"},inplace=True)
        df_returns = pd.read_csv(report_return_path,low_memory=False)
        df_returns.rename(columns={"ORDER_ID":"Номер по данным клиента"},inplace=True)
        extract_path = Path("unzipped_report")
        if extract_path.exists():
            shutil.rmtree(extract_path)
        return df_payments,df_returns


def get_list_shops():
    url = "https://api.partner.market.yandex.ru/campaigns"
    params = {
        "page":1,
        "pageSize":20
    }
    response = session.get(url = url,params=params,timeout=120)
    data = response.json().get("campaigns",[])
    print(data)

def save_to_excel(df):
    sheet_name = "2025"
    result = reorder_columns(df)

    fill_pink = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # светло-розовый
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # светло-жёлтый
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # светло-зелёный

    pink_cols = {"Номер", "Дата", "Отчетный месяц", "Номер по данным клиента",
                "Сумма из 1С", "Сумма итого", "Проверка", "Комментарий"}
    yellow_cols_keywords = {"Сумма из ЛК ЯМ", "Статус заказа", "Баллы"}
    locale.setlocale(locale.LC_TIME, "en_US.UTF-8")
    file_path = resource_path("TEST.xlsx")
    mode = "a" if Path(file_path).exists() else "w"

    writer_kwargs = {
        "engine": "openpyxl",
        "mode": mode
    }
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"
    with pd.ExcelWriter(file_path, **writer_kwargs) as writer:
        result.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 20
        for _, cell in enumerate(ws[1], start=1):
            col_name = str(cell.value)
            if col_name in pink_cols:
                cell.fill = fill_pink
            elif any(keyword in col_name for keyword in yellow_cols_keywords):
                cell.fill = fill_yellow

            cell.alignment = Alignment(wrap_text=True,horizontal="center",vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for idx, cell in enumerate(row, start=1):
                col_name = ws.cell(row=1, column=idx).value
                if col_name == "Номер по данным клиента" or col_name =="Отчетный месяц":
                    continue
                if isinstance(cell.value, (float)):
                    if cell.value.is_integer():
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = "#,##0.##"
                elif isinstance(cell.value, (int)):
                    cell.number_format = '#,##0'
                elif cell.value == "перевели меньше":
                    cell.fill = fill_green

def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        "Номер", "Дата", "Отчетный месяц", "Номер по данным клиента",
        "Сумма из 1С", "Сумма из ЛК ЯМ", "Статус заказа"
    ]
    summary_cols = ["Сумма итого", "Проверка", "Комментарий"]

    def parse_col(col: str):
        match = re.match(r"(\d{4})\s*([А-Яа-я]+)", col)
        if not match:
            return (9999, 13, 99)

        year = int(match.group(1))
        month_name = match.group(2)
        month_num = list(calendar.month_name).index(month_name.capitalize()) if month_name.capitalize() in calendar.month_name else 13

        if "Начисления" in col:
            order = 1
        elif "Возвраты" in col:
            order = 2
        elif "Баллы" in col:
            order = 3
        else:
            order = 99

        return (year, month_num, order)

    sum_cols = [c for c in df.columns if any(x in c for x in ["Начисления", "Возвраты", "Баллы"])]
    sum_cols = sorted(sum_cols, key=parse_col)

    all_cols = [c for c in base_cols if c in df.columns] + sum_cols + [c for c in summary_cols if c in df.columns]

    return df.reindex(columns=all_cols).apply(lambda col: col.fillna(0) if col.name != "Комментарий" else col)


def create_united_netting_report(bank_order_id,date_of_report):
    print(bank_order_id)
    print(date_of_report)
    url_create = "https://api.partner.market.yandex.ru/reports/united-netting/generate"
    params = {"format":"CSV","language":"RU"}
    payload = {
        "businessId": BUSINESSID,
        "bankOrderId": bank_order_id,
        "bankOrderDateTime":date_of_report
    }
    response = session.post(url=url_create,params=params,json = payload)
    if response.status_code == 200:
        data = response.json()
        report_id = data.get("result",{}).get("reportId")
        print("Отчёт YandexMarket успешно создан. Информация:")
        print(f"ID отчета: {report_id}")
        return report_id
    else:
        print(f"Ошибка запроса: {response.status_code}")
        print(response.text)

def main():
    print("Начинаем обработку Yandex Market")
    month_input = input("Введите месяц (например: апрель): ").strip().lower()
    year_input = 2025

    first_day = datetime.strptime(f"{month_input} {year_input}", "%B %Y").date()
    last_day_num = calendar.monthrange(year_input, first_day.month)[1]
    last_day = date(year_input, first_day.month, last_day_num)

    date_from = last_day - relativedelta(months=3)
    date_to = last_day

    folder_path = resource_path("DATA")
    month = date_to.strftime("%B")

    df_onec = pd.read_excel(f"{folder_path}\\{month}\\Заказы 1С {month} ЯМ.xlsx") \
        .dropna(subset=["Номер"]).drop(columns=["Комментарий"])
    df_onec["Дата"] = pd.to_datetime(df_onec["Дата"], format="%d.%m.%Y %H:%M:%S")
    df_onec["Отчетный месяц"] = df_onec["Дата"].dt.month
    df_onec.rename(columns={"Сумма": "Сумма из 1С"}, inplace=True)
    df_onec["Сумма из 1С"] = df_onec["Сумма из 1С"].round()

    file_path = resource_path("TEST.xlsx")
    key = "Номер по данным клиента"
    if os.path.exists(file_path):
        existing_df = pd.read_excel(file_path)
    else:
        existing_df = pd.DataFrame()

    combined = pd.concat([existing_df, df_onec], ignore_index=True)

    print(f"Создается отчет по {date_from}  -  {date_to}")

    report_id = create_yandex_report(date_from=str(date_from), date_to=str(date_to))
    df = get_yandex_report(report_id=report_id)
    df.rename(columns={
        'PARTNER_ORDER_ID': "Номер",
        "OFFER_STATUS": "Статус заказа",
        "ORDER_ID": key,
        "NETTING_AMOUNT": "Сумма баллов"
    }, inplace=True)
    df["Сумма из ЛК ЯМ"] = df["DELIVERED_OR_RETURNED"] * df["BILLING_PRICE"]
    df = df.groupby(key).agg({
        "Сумма из ЛК ЯМ": "sum",
        "Статус заказа": "first",
        "Сумма баллов": "sum",
        "REFUND_BUYER_PAYMENT_BANK_ORDER_ID": "first",
        "REFUND_BUYER_PAYMENT_BANK_ORDER_DATE": "first"
    }).reset_index()

    combined = combined.merge(df, on=key, how="left", suffixes=("", "_new"))
    combined.dropna(subset="Номер", inplace=True)

    for col in df.columns:
        if col != key:
            col_new = col + "_new"
            if col_new in combined.columns:
                combined[col] = combined[col_new].where(combined[col_new].notna(), combined[col])
                combined.drop(columns=[col_new], inplace=True)

    bank_pairs_filtered = (
        combined[["REFUND_BUYER_PAYMENT_BANK_ORDER_ID", "REFUND_BUYER_PAYMENT_BANK_ORDER_DATE"]]
        .drop_duplicates(subset=["REFUND_BUYER_PAYMENT_BANK_ORDER_ID"])
        .dropna()
    )

    for _, row_filtered in bank_pairs_filtered.iterrows():
        bank_id = int(row_filtered["REFUND_BUYER_PAYMENT_BANK_ORDER_ID"])
        bank_date_val = pd.to_datetime(row_filtered["REFUND_BUYER_PAYMENT_BANK_ORDER_DATE"],format = "%d.%m.%Y")
        bank_date_str = bank_date_val.strftime("%Y-%m-%dT%H:%M:%SZ")
        json_file = resource_path("reports.json")

        if json_file.exists():
            with open(json_file, "r", encoding="utf-8") as f:
                reports = json.load(f)
        else:
            reports = {}

        bank_id_str = str(bank_id)

        if bank_id_str in reports:
            report_id = reports[bank_id_str]
            print(f"Нашли в JSON: {bank_id} -> {report_id}")
        else:
            report_id = create_united_netting_report(
                bank_order_id=bank_id,
                date_of_report=bank_date_str
            )
            reports[bank_id_str] = report_id
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(reports, f, ensure_ascii=False, indent=4)
            print(f"Добавили в JSON: {bank_id} -> {report_id}")

        previous_month_date = bank_date_val - relativedelta(months = 1)
        month_name = previous_month_date.strftime("%B")
        df_payments, df_returns = get_yandex_report(report_id=report_id, mask=False)

        df_payments = df_payments.groupby("Номер по данным клиента").agg({"TRANSACTION_SUM": "sum"}).reset_index()
        df_returns = df_returns.groupby("Номер по данным клиента").agg({"TRANSACTION_SUM": "sum"}).reset_index()
        df_payments.rename(columns={"TRANSACTION_SUM": "TRANSACTION_SUM_pay"}, inplace=True)
        df_returns.rename(columns={"TRANSACTION_SUM": "TRANSACTION_SUM_ret"}, inplace=True)

        combined = combined.merge(df_payments, how="left", on="Номер по данным клиента")
        combined = combined.merge(df_returns, how="left", on="Номер по данным клиента")

        combined[f"{previous_month_date.year} {month_name} Начисления"] = combined["TRANSACTION_SUM_pay"].fillna(0)
        combined[f"{previous_month_date.year} {month_name} Возвраты"] = combined["TRANSACTION_SUM_ret"].fillna(0)

        combined.drop(columns=["TRANSACTION_SUM_pay", "TRANSACTION_SUM_ret"], inplace=True, errors="ignore")
    combined = combined.drop_duplicates(subset="Номер")
    combined[f"{date_to.year} {month_input.capitalize()} Баллы"] = combined["Сумма баллов"].where(
        combined["Дата"].dt.month.map(lambda m: calendar.month_name[m]) == month_input.capitalize()
    )
    cols = combined.columns.tolist()
    if "Сумма итого" not in combined.columns:
        combined["Сумма итого"] = None
    russian_months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    all_results = []
    for month in russian_months:
        sum_cols = [col for col in cols if month in col]
        all_results.extend(sum_cols)
    combined[all_results] = combined[all_results].apply(pd.to_numeric, errors="coerce").fillna(0)
    combined["Сумма итого"] = combined[all_results].sum(axis = 1)
    def check_row(row):
        if row["Сумма итого"] == row["Сумма из ЛК ЯМ"]: return "ок"
        elif row["Сумма итого"] > row["Сумма из ЛК ЯМ"]: return "перевели больше"
        elif row["Сумма итого"] < row["Сумма из ЛК ЯМ"]: return "перевели меньше"
        else: return "-"
    combined["Проверка"] = combined.apply(check_row, axis=1)

    manager_counts = combined.groupby("Номер по данным клиента")["Автор"].transform(lambda x: x.isin(MENEGERS).sum())

    combined.loc[
        (combined["Автор"] == "reglament_yandex_market") & (manager_counts > 0), "Статус заказа"] = "Отменен"

    if "Комментарий" not in combined.columns:
        combined["Комментарий"] = ''
    print("Сохранение...")
    save_to_excel(combined)

if __name__ == "__main__":
    try:
        main()
        input("Введите ENTER для завершения")
    except Exception:
        traceback.print_exc()
        input("Скопируйте ошибку для отправки и нажмите ENTER")